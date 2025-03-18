import pandas as pd
import numpy as np
import datetime
import openpyxl

from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.utils import get_column_letter

from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from io import BytesIO

file_path = 'Expense_Tracker_Template.xlsx'
output_file_path = 'output.xlsx'

wb = openpyxl.Workbook()

header_style = NamedStyle(name="header_style")
header_style.font = Font(bold=True, color="FFFFFF")
header_style.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_style.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
header_style.alignment = Alignment(horizontal="center")

wb.add_named_style(header_style)

wb.save(output_file_path)

budget_df = pd.read_excel(file_path, sheet_name='Budget')
expenses_df = pd.read_excel(file_path, sheet_name='Expenses')

grouped = expenses_df.groupby([expenses_df['Date'].dt.year, expenses_df['Date'].dt.month])
sorted_groups = sorted(
    [(year, month, df) for (year, month), df in grouped],
    key=lambda x: (x[0], x[1]),  
    reverse=True 
)

def append_df_data(df):
    header = df[0]
    ws.append(header)

    for col in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=col).style = header_style
    
    start_row = ws.max_row

    data = df[1:]
    for row in data:
        ws.append(row)

    for row in range(start_row, ws.max_row + 1):
        for i in range(len(header)):
            if header[i] in ["Total Monthly Spending", "Average Daily Spending", "Total Spent", "Budget", "Total Spending"]:
                ws.cell(row=row, column=i + 1).number_format = FORMAT_CURRENCY_USD_SIMPLE
            if header[i] in ["% of Total"]:
                ws.cell(row=row, column=i + 1).number_format = '0.0%'
            if header[i] in ["Budget vs. Actual"]:
                if "On Track" in ws.cell(row=row, column=i + 1).value:
                    ws.cell(row=row, column=i + 1).style = 'Good'
                if "Over Budget" in ws.cell(row=row, column=i + 1).value:
                    ws.cell(row=row, column=i + 1).style = "Bad"
    ws.append([])

for year, month, df in sorted_groups:
    index = sorted_groups.index((year, month, df))
    sliced = sorted_groups[index:] 

    month_name = pd.to_datetime(f"{year}-{month}-01").strftime('%b')
    name = f"{year}, {month_name}"

    current_grouped_df = df.groupby('Category')['Amount'].sum()

    comparison_df = pd.merge(budget_df, current_grouped_df, on='Category', how='outer')
    if comparison_df['Amount'].isnull().values.any():
        comparison_df['Amount'] = comparison_df['Amount'].fillna(0)
    comparison_df['Difference'] = comparison_df['Monthly Budget'] - comparison_df['Amount']
    comparison_df['Status'] = comparison_df['Difference'].apply(
        lambda x: f'${-x} Over Budget' if x < 0 else 'On Track'
    )

    monthly_spending = df['Amount'].sum()
    avg_daily = df['Amount'].sum() / df['Date'].dt.day.max()
    most_expensive_categ = f'{current_grouped_df.idxmax()} (${current_grouped_df.max()})'
    budget_status = f'${-comparison_df['Difference'].sum()} Over Budget' if comparison_df['Difference'].sum() < 0 else 'On Track'
    
    key_metrics = [
        ["Total Monthly Spending", "Average Daily Spending", "Most Expensive Category", "Budget Status"],
        [monthly_spending, avg_daily, most_expensive_categ, budget_status]
    ]

    categories = comparison_df['Category'].values
    total_spent = comparison_df['Amount'].values
    percent_of_total = [round((num / monthly_spending) * 100, 2) for num in comparison_df['Amount'].values]
    budget = comparison_df['Monthly Budget'].values
    budget_vs_actual = comparison_df['Status'].values

    category_breakdown = [
        ['Category', 'Total Spent', '% of Total', 'Budget', 'Budget vs. Actual']
    ] + [
        [category, spent, (spent / monthly_spending), budg, status]
        for category, spent, budg, status in zip(categories, total_spent, budget, budget_vs_actual)
    ]

    monthly_trends = [["Month", "Total Spending", "Most Spent Category"]]
    specified_year = year

    for year, month, df in sliced:
        if year == specified_year:
            category_totals = df.groupby('Category')['Amount'].sum()

            month_name = pd.to_datetime(f"2000-{month}-01").strftime('%B')
            total_spending = df['Amount'].sum()
            most_spent = f'{category_totals.idxmax()} (${category_totals.max()})'
            monthly_trends.append([month_name, total_spending, most_spent])

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.pie(comparison_df['Amount'], labels=comparison_df['Category'], autopct='%1.1f%%')

    chart_image = BytesIO()
    plt.savefig(chart_image, format='png', bbox_inches='tight') 
    plt.close() 

    chart_image.seek(0)
    pie_chart = Image(chart_image)

    with pd.ExcelWriter(output_file_path, mode='a', engine='openpyxl') as writer:
        if name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(name)
            sheet = writer.book.worksheets[idx]
            writer.book.remove(sheet) 

        pd.DataFrame().to_excel(writer, sheet_name=name, index=False)
        ws = writer.book[name]

        append_df_data(key_metrics)

        append_df_data(category_breakdown)

        append_df_data(monthly_trends)

        dimension = ws.calculate_dimension()
        max_col = dimension.split(":")[1]
        max_col_letter = max_col[0] 
        max_col_number = ord(max_col_letter.upper()) - ord("A") + 1  
        new_col_number = max_col_number + 2 
        new_col_letter = chr(ord("A") + new_col_number - 1) 

        ws.add_image(pie_chart, f'{new_col_letter}1') 

        for col in range(1, max_col_number + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(output_file_path)
        pass